#!/usr/bin/env python3
"""Generate Excel spec sheet from docs markdown files.

Usage:
    pip install openpyxl
    python docs/gen-spec-excel.py
Output: docs/spec.xlsx
"""
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

DOCS_DIR = Path(__file__).parent
OUTPUT = DOCS_DIR / "spec.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
SUBHEADER_FONT = Font(bold=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def parse_md_tables(md_text):
    """Extract all markdown tables as list of (headers, rows)."""
    tables = []
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("|") and i + 1 < len(lines):
            sep = lines[i + 1].strip()
            if re.match(r"^\|[-| :]+\|$", sep):
                headers = [c.strip() for c in line.strip("|").split("|")]
                rows = []
                j = i + 2
                while j < len(lines) and lines[j].strip().startswith("|"):
                    row = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                    rows.append(row)
                    j += 1
                tables.append((headers, rows))
                i = j
                continue
        i += 1
    return tables


def write_sheet(wb, sheet_name, md_file):
    ws = wb.create_sheet(title=sheet_name)
    text = md_file.read_text(encoding="utf-8")
    row_idx = 1

    # Write title
    ws.cell(row=row_idx, column=1, value=md_file.stem.replace("-", " ").title())
    ws.cell(row=row_idx, column=1).font = Font(size=14, bold=True)
    row_idx += 2

    # Write each table
    for headers, rows in parse_md_tables(text):
        # Header row
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        # Data rows
        for r, row in enumerate(rows):
            for col, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if r % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="DEEAF1")
            row_idx += 1
        row_idx += 1  # blank line between tables

    auto_width(ws)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_sheet(wb, "Functional Spec", DOCS_DIR / "functional-spec.md")
    write_sheet(wb, "API Spec", DOCS_DIR / "api-spec.md")
    write_sheet(wb, "Canary", DOCS_DIR / "canary.md")

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
